import warnings
from copy import deepcopy
from typing import Union, Sequence, overload, List, Generic, TypeVar, Optional, Dict, Iterable

import win32com
from myfile import get_elements_from_line
from openpyxl import *
from openpyxl.cell.cell import Cell
from statistics_nyh import absmax
from win32com.client import constants as c  # 旨在直接使用VBA常数

from GoodToolPython.mybaseclasses.tools import is_sequence_with_specified_type
import unittest



def bisection_method(sorted_list:list,
                     goal,
                     func:callable):
    """
    二分法查找
    @param sorted_list: 已经按升序排列好的列表，或者可以进行下标操作的对象 程序不会检查这个条件，请确保这个参数已经按升序排好 否则可能返回错误结果
    @param goal: 目标
    @param func: 函数 其输入参数为sorted_list的元素
    @return: 如果成功找到，返回id,lbound,ubound  id是基于0 如果goal完全找到，lb和ub就有两种可行的值
            如果找不到,返回None,lbound,ubound#没找到 返回最接近的两个id
    """
    #检查goal是否在上下界之间
    assert callable(func),'func必须为函数'
    assert len(sorted_list)>=2,'sorted_list必须至少两个元素'
    if goal==func(sorted_list[0]):
        return 0,0,1
    elif goal==func(sorted_list[-1]):
        return len(sorted_list)-1, len(sorted_list)-2, len(sorted_list)-1
    else:
        assert func(sorted_list[0])<goal<func(sorted_list[-1]),'goal超出sorted_list的上下界'
    lbound = 0
    ubound = len(sorted_list) - 1  # 二分法的上下界
    while True:
        #检查是否结束
        if ubound-lbound<=1:
            return None,lbound,ubound#没找到 返回最接近的两个id
        midone_id=round((lbound+ubound)/2.0)
        midone=func(sorted_list[midone_id])
        if goal==midone:
            return midone_id,midone_id-1,midone_id
        elif goal<midone:
            #在前半段
            ubound=midone_id
        else:#在后半段
            lbound=midone_id

def myunique(lst, return_firstid=False):
    """
    去除重复元素
    自编
    按照首次出现的位置排序
    @param lst:
    @param return_firstid: 是否返回第一次出现的id
    @return:
    """
    lst1=list(set(lst))
    lst1.sort(key=lst.index)
    if return_firstid:
        firstid=[lst.index(x) for x in lst1]
        return lst1,firstid
    else:
        return lst1


# T_Model=TypeVar('T_Model')


# @overload
# class FlatDataModel:
#     pass


class DataUnit:
    """
    数据单元 指一行
    """

    @staticmethod
    def make(vn: Sequence[str], row: Sequence[Cell], model: 'FlatDataModel',mergedcells:dict):
        """
        这个函数用于从excel文件中生成数据单元
        :param vn:
        :param row:
        :param model:
        :@param mergecells:传递合并单元格信息
        :return:
        """
        self = DataUnit()
        # assert is_sequence_with_specified_type(row, Cell), 'row必须是Cell组成的列表'
        assert is_sequence_with_specified_type(vn, str), 'vn必须是str组成的序列'
        assert isinstance(model, FlatDataModel), 'model必须是FlatDataModel对象'
        assert len(vn) == len(row), 'vn与row必须大小一致'
        self.data = {}  # type: Union[int,str,float]#以字典的形式保存所有变量
        for name, value in zip(vn, row):
            if "MergedCell" in str(type(value)):#若是合并单元格
                self.data[name]=value.parent.cell(mergedcells[(value.row,value.column,)][0],mergedcells[(value.row,value.column,)][1]).value
            else:#普通单元格
                self.data[name] = value.value
        self.model = model  # type:FlatDataModel   #保存模型
        return self

    def __init__(self, model=None):
        self.data = {}  #以字典的形式保存所有变量
        if model is not None:
            assert isinstance(model, FlatDataModel), '参数错误'
        self.model = model  # type:FlatDataModel   #保存模型

    def __len__(self):
        # 返回变量的个数
        return len(self.data.keys())

    def __getitem__(self, item):
        if is_sequence_with_specified_type(item, str):
            # 返回多个变量值
            r = []
            for name in item:
                r.append(self.data[name])
            return tuple(r)
        else:
            return self.data.__getitem__(item)

    def __iter__(self):
        # 迭代器 依次返回该单元的所有变量值
        self.__iter_id = 0
        return self

    def __next__(self):
        id = self.__iter_id
        if id == len(self):
            raise StopIteration
        self.__iter_id += 1
        return self.data[self.model.vn[id]]


class FlatDataModel:
    """
    平面数据模型 excel文件中的数据 一行代表一个数据单元
    所有的序号都是从0数起
    """

    @staticmethod
    def load_from_list(vn_lst, data_lst) -> 'FlatDataModel':
        """
        从列表中生成
        :param vn_lst: 变量名列表
        :param data_lst: 数据列表 2维列表 每一个内层列表代表一个数据点
        :return:
        """
        assert is_sequence_with_specified_type(vn_lst, str), '必须为str列表'
        assert isinstance(data_lst,Iterable),"数据必须可迭代"
        # assert is_sequence_with_specified_type(data_lst, list) or is_sequence_with_specified_type(data_lst, tuple), '必须为列表组成的列表'
        fdm = FlatDataModel()
        fdm.vn = vn_lst

        for row in data_lst:
            u = DataUnit(fdm)
            assert len(row) == len(vn_lst), '数据与变量个数必须一致'
            for nm, v in zip(vn_lst, row):
                u.data[nm] = v
            fdm.units.append(u)
        return fdm

    @staticmethod
    @overload
    def load_from_excel_file(fullname) -> 'FlatDataModel':
        pass

    @staticmethod
    @overload
    def load_from_excel_file(fullname,
                             sheetname: str = None,
                             row_variable_name: int = 0,
                             row_caption: List[int] = None,
                             row_data_start: int = None) -> 'FlatDataModel':
        pass

    @staticmethod
    def load_from_excel_file(fullname, sheetname=None, row_variable_name=0, row_caption=None,
                             row_data_start=None) -> 'FlatDataModel':
        """
        从excel'文件中载入平面数据模型
        允许单元格有竖向的合并，合并的单元格看做是后续合并行与第一行有相同数据
         不允许有横向的合并，但程序未做检查
        :param fullname:
        :param sheetname:
        :param row_variable_name: 变量名所在行 0-based
        :param row_caption: 注释行 可以是列表，代表多行注释 默认为空
        :param row_data_start: 数据起始行 默认为变量名行下一行
        :return:
        """
        self = FlatDataModel()
        workbook = load_workbook(fullname, data_only=True)
        assert len(workbook.sheetnames) > 0, '此表无工作簿'
        if sheetname is None:
            sheetname = workbook.sheetnames[0]  # 默认为第一个工作簿
        else:
            assert sheetname in workbook.sheetnames, '工作簿不存在'
        pass
        worksheet = workbook[sheetname]
        rows = list(worksheet.rows)  # 将工作簿中所有数据转化为以行为单位的列表

        #处理合并单元格 允许单元格有竖向的合并 不允许有横向的合并
        mergecells={}
        if worksheet.merged_cells:
            for item in worksheet.merged_cells:#对每一处合并单元格进行循环
                for row in range(item.min_row, item.max_row+1):
                    for col in range(item.min_col, item.max_col+1):
                        #用位置坐标作为key，value为首单元格坐标
                        mergecells[(row,col)]=(item.min_row,item.min_col,)
        # 读变量名
        self.vn = [cell.value for cell in rows[row_variable_name]]  # type: List[str]
        assert is_sequence_with_specified_type(self.vn, str), '变量名读取失败'
        #检查变量名是否重复
        t=set(self.vn)
        if len(t)<len(self.vn):
            raise Exception("变量名重复")
        # 读注释行
        self.caption = []  # type:List[List[str]]
        if row_caption is not None:
            if isinstance(row_caption, int):
                row_caption = [row_caption]
            for row_id in row_caption:
                tmp = [cell.value for cell in rows[row_id]]
                self.caption.append(tmp)


        # 读数据
        self.units = []  # type:list[DataUnit]
        if row_data_start is None:
            if row_caption is None:
                row_data_start = row_variable_name + 1
            else:
                row_data_start = row_caption[-1] + 1
        for row in rows[row_data_start:]:
            unit = DataUnit.make(self.vn, row, self,mergecells)
            self.units.append(unit)
        return self

    # def __init__(self,fullname, sheetname=None, row_variable_name=0, row_caption=None, row_data_start=None):
    #     workbook = load_workbook(fullname, data_only=True)
    #     assert len(workbook.sheetnames) > 0, '此表无工作簿'
    #     if sheetname is None:
    #         sheetname = workbook.sheetnames[0]  # 默认为第一个工作簿
    #     else:
    #         assert sheetname in workbook.sheetnames, '工作簿不存在'
    #     pass
    #     worksheet = workbook[sheetname]
    #     rows = list(worksheet.rows)  # 将工作簿中所有数据转化为以行为单位的列表
    #
    #     # 读变量名
    #     self.vn = [cell.value for cell in rows[row_variable_name]]# type: List[str]
    #     assert is_sequence_with_specified_type(self.vn, str), '变量名读取失败'
    #
    #     # 读注释行
    #     self.caption = []  # type:List[List[str]]
    #     if row_caption is not None:
    #         if isinstance(row_caption,int):
    #             row_caption=[row_caption]
    #         for row_id in row_caption:
    #             tmp=[cell.value for cell in rows[row_id]]
    #             self.caption.append(tmp)
    #
    #     # 读数据
    #     self.units=[]# type:list[DataUnit]
    #     if row_data_start is None:
    #         if row_caption is None:
    #             row_data_start=row_variable_name+1
    #         else:
    #             row_data_start=row_caption[-1]+1
    #     for row in rows[row_data_start:]:
    #         unit = DataUnit(self.vn, row,self)
    #         self.units.append(unit)

    def __init__(self):
        self.vn = []  # type:list[str]
        self.units = []  # type:list[DataUnit]
        self.caption = []  # type:list[list[str]]

    def __len__(self):
        # 返回units的个数
        return len(self.units)

    def __getitem__(self, item) -> Union[DataUnit, list]:
        """
        获取部分数据点（切片） 或者获取所有数据点的某个变量信息
        :param item:
        :return:
        """
        # if isinstance(item,str):
        #     item=[item]
        if isinstance(item,list) or isinstance(item,str):  # 获取所有数据点的某个变量信息
            t=[x[item] for x in self.units]
            return t
            # assert item in self.vn, '该变量名不存在'
            # r = []
            # for u in self:
            #     r.append(u[item])
            # return r
        else:
            return self.units.__getitem__(item)

    def __setitem__(self, key, value):
        """
        增加一个字段
        如果该字段已存在会打印警告信息
        @param key:
        @param value: 函数 操作对象为dataunit 如：fdm['新字段名']=lambda x:-1*x['已有字段名']
                        list 长度与自身数据点个数相等
        @return:
        """
        def script_func():
            for u in self:
                u.data[key] = value(u)

        def script_lst():
            assert len(value)==len(self),"列表长度与自身数据点个数不一致。"
            for i,u in enumerate(self):
                u.data[key]=value[i]

        if key in self.vn:
            print("字段(%s)被覆盖" % key)
        else:
            self.vn.append(key)  # 更新vn
        if callable(value):
            script_func()
        elif isinstance(value,list):
            script_lst()
        else:
            raise TypeError("参数错误")

    def save(self, fullname, sheetname=None):
        # 保存到excel文件中
        wb = Workbook()
        self.__add_to_workbook(wb, 'Sheet')

        # 保存文件
        wb.save(fullname)

    def save_to_txt(self,fullname,txt_func,write_vn=True):
        with open(fullname, 'w') as file_object:
            if write_vn:#写入vn
                file_object.write(",".join(self.vn))
                file_object.write("\n")
            for u in self:
                file_object.write(txt_func(u))
        pass
    def __iter__(self):
        # 依次返回每一个unit
        return self.units.__iter__()

    def flhz(self,
             classify_names: Union[str, List[str]],
             statistics_func: List[List],
             flag_write_statistics_func=False) -> 'FlatDataModel':
        """
        分类汇总
        :param classify_names:
        :param statistics_func: 列表组成的列表 二级嵌套列表 子列表形如[字段名,统计函数]或[字段名,统计函数,新字段名]
        :param flag_write_statistics_func:
        :return:
        """
        # 参数预处理
        if isinstance(classify_names, str):
            classify_names = [classify_names]
        assert is_sequence_with_specified_type(classify_names, str), 'is_sequence_with_specified_type参数错误'
        assert isinstance(statistics_func, list), 'statistics_func必须为list'
        for i, line in enumerate(statistics_func):
            assert isinstance(line, list) and \
                   line[0] in self.vn and \
                   callable(line[1]), 'statistics_func中的元素必须为[str,函数,str]或者[str,函数]'
            if len(line) == 2:
                line.append(line[0])
        # assert isinstance(statistics_func,dict),'statistics_func必须为字典'
        #
        # for s_name,val in statistics_func.items():
        #     assert s_name in self.vn,'键必须为变量名'
        #     if callable(val):
        #         # val=[val,s_name]
        #         statistics_func[s_name]=[val,s_name]
        #     else:
        #         assert isinstance(val,list) and callable(val[0]) and isinstance(val[1],str),'值必须为[函数,str]'

        model_copy = deepcopy(self)  # 复制
        return_model = FlatDataModel()  # 返回值
        # 把classify_names做成函数

        # 排序
        model_copy.units.sort(key=lambda x: x[classify_names])

        # 按classify_names分类
        current_classify_value = model_copy[0][classify_names]
        bunch = []  # type:list[DataUnit]
        bunchs = []  # type:list[list[DataUnit]]
        for unit in model_copy:
            this_classify_value = unit[classify_names]
            if this_classify_value == current_classify_value:
                bunch.append(unit)
            else:
                bunchs.append(bunch)
                bunch = [unit]
                current_classify_value = this_classify_value
        if len(bunch) != 0:
            bunchs.append(bunch)

        # 按变量名提取值
        for bunch in bunchs:
            # 对每一个bunch进行提值 统计 生成一个单元
            unit = DataUnit(return_model)
            # 添加classify变量
            for c_name in classify_names:
                unit.data[c_name] = bunch[0][c_name]
            # 添加统计变量
            for line in statistics_func:
                # for s_name in [x[0] for x in statistics_func]:
                s_name = line[0]
                func = line[1]
                newname = line[2]
                value_list = [x[s_name] for x in bunch]
                statistics_value = func(value_list)
                unit.data[newname] = statistics_value
            return_model.units.append(unit)

        # 处理vn
        return_model.vn = classify_names + [line[2] for line in statistics_func]

        return return_model

    def append_to_file(self, fullname, sheetname):
        workbook = load_workbook(fullname, data_only=True)
        self.__add_to_workbook(workbook, sheetname)
        workbook.save(fullname)

    def __add_to_workbook(self, wb, sheetname):
        """
        将自身以sheet的形式添加到wb中，
        :param wb:
        :return:
        """
        assert isinstance(wb, Workbook), 'wb为workbook对象'
        if sheetname in wb.sheetnames:
            # raise Exception("该工作簿已存在")
            warnings.warn('原有工作簿被删除')
            wb.remove(wb[sheetname])
            sheet = wb.create_sheet(sheetname)
        else:
            # 创建
            assert isinstance(sheetname, str), 'sheetname类型必须为str'
            sheet = wb.create_sheet(sheetname)

        # 写入标题行
        for col in range(1, 1 + len(self.vn)):
            sheet.cell(row=1, column=col).value = self.vn[col - 1]

        row_rd = 2
        col_rd = 1  # 下一个可以写入的位置
        # 写入注释行
        for rowin in self.caption:
            for cellin in rowin:
                sheet.cell(row=row_rd, column=col_rd).value = cellin
                col_rd += 1
            row_rd += 1
            col_rd = 1

        # 写入数据
        for unit in self:
            for v in unit:
                sheet.cell(row=row_rd, column=col_rd).value = v
                col_rd += 1
            row_rd += 1
            col_rd = 1

    def add_variables_from_other_model(self,
                                       other: 'FlatDataModel',
                                       link_variable: str,
                                       add_variable: Union[str, List[str]] = None,
                                       continue_on_not_found=True) -> None:
        """

        @param other:
        @param link_variable:
        @param add_variable: 为None时，默认添加other中除link_var的全部变量
        @param continue_on_not_found: 如果没找到 是否继续找下一个
        @return:
        """
        # 参数处理
        assert isinstance(other, FlatDataModel), 'other必须为FlatDataModel对象'
        assert isinstance(link_variable, str) and \
               link_variable in self.vn and \
               link_variable in other.vn, \
            'link_variable必须是两个模型中均存在的变量名'
        if add_variable is None:
            # 默认添加全部, 除了连接变量
            add_variable = [x for x in other.vn if x != link_variable]
        elif isinstance(add_variable, str):
            add_variable = [add_variable]
        else:
            assert is_sequence_with_specified_type(add_variable, str), 'add_variable为字符串列表'

        # 开始添加
        for unit in self:
            link_variable_value = unit[link_variable]
            expr = lambda x: x[link_variable].__eq__(link_variable_value)
            cor_unit = other.find(expr)  # 找到此模型的unit对应的在另外一个unit
            if cor_unit is None:  # 没找到
                if continue_on_not_found is False:
                    raise Exception("另外一个模型中未找到对应的unit")
                else:
                    continue
            for name in add_variable:
                unit.data[name] = cor_unit[name]

        # 处理数变量名
        self.vn += add_variable

    def find(self, epxr, flag_find_all=False) -> Union[DataUnit, List[DataUnit]]:
        """
        查找满足epxr的unit
        :param epxr: 真假判断函数
        :param flag_find_all: 是否查找全部
        :return: 找不到的时候 返回None []
        """
        assert callable(epxr), 'expr必须为函数'
        r = []
        for unit in self:
            if epxr(unit) is True:
                if flag_find_all is False:
                    return unit
                else:
                    r.append(unit)
        if flag_find_all is False:
            return None
        else:
            return r

    def show_in_excel(self):
        """
        在excel中显示 此方法不会阻塞
        :return:
        """
        """
        这个函数有时候会报错 。。。has no attribute 'CLSIDToClassMap'
        解决办法：
        运行：
        from win32com.client.gencache import EnsureDispatch
        import sys
        xl = EnsureDispatch("Word.Application")
        print(sys.modules[xl.__module__].__file__)
        以上代码给出了临时文件路径，删除gen_py下所有长数字文件夹即可解决
        """
        xl_app = win32com.client.gencache.EnsureDispatch("Excel.Application")  # 若想引用常数的话使用此法调用Excel
        xl_app.Visible = False  # 是否显示Excel文件
        wb = xl_app.Workbooks.Add()
        sht = wb.Worksheets(1)
        row_rd, col_rd = 1, 1  # 记录下一个写入的位置
        for name in self.vn:
            sht.Cells(row_rd, col_rd).Value = name
            col_rd += 1
        row_rd = 2
        col_rd = 1
        for rowin in self.caption:
            for cellin in rowin:
                sht.Cells(row_rd, col_rd).Value = cellin
                col_rd += 1
            row_rd += 1
            col_rd = 1
        for unit in self:
            for v in unit:
                sht.Cells(row_rd, col_rd).Value = v.__str__()
                col_rd += 1
            row_rd += 1
            col_rd = 1
        xl_app.Visible = True

    def __str__(self):
        s1 = '%d个变量' % len(self.vn)
        s2 = self.vn.__str__()
        s3 = '%d个数据单元' % len(self.units)
        return "%s\n%s\n%s" % (s1, s2, s3)

    def to_list(self, write_variable_names=False):
        """
        将所有数据转化为2维列表并返回 一行代表一个数据点
        :return:
        """
        t = [[x2 for x2 in x1] for x1 in self.units]
        if write_variable_names is True:
            t.insert(0, deepcopy(self.vn))
        return t
        pass

    def __delete_variable(self, nm):
        assert nm in self.vn, '变量名不存在'
        for u in self:
            del u.data[nm]
        self.vn.remove(nm)

    def delete_variable(self, variable_name: Union[str, List[str]]) -> None:
        """
        删除某些字段名
        :param variable_name: 字符串或字符串列表
        :return:
        """
        if isinstance(variable_name, str):
            variable_name = [variable_name]
        assert is_sequence_with_specified_type(variable_name, str), '必须为字符串或字符串列表'
        for nm in variable_name:
            self.__delete_variable(nm)

    def narrow(self, vn_lst: List[str], reserve=True) -> 'FlatDataModel':
        """
        瘦身为一个新的平面数据模型
        不会改变自身 而是复制一个新的数据模型进行操作
        :param vn_lst: 变量名字符串列表
        :param reserve: 指定是要保留的 还是要删除的
        :return:
        """
        fdm = deepcopy(self)
        if True == reserve:  # 处理要删除的变量名
            nms_del = deepcopy(self.vn)
            for i in vn_lst:
                nms_del.remove(i)
        else:
            nms_del = vn_lst
        fdm.delete_variable(nms_del)  # 删除
        return fdm

    @staticmethod
    def load_from_file(pathname, vn_style: Union[str, list] = 'numbers', omitlines=0, separator: str = '',
                       column_expected: int = None) -> 'FlatDataModel':
        """
        从文本文件中载入
        @param pathname:
        @param vn_style: ‘numbers’ 使用自动递增的数据vn0 vn1 vn2作为vn 其长度可以通过column_expected指定；当column_expected=0时 程序自动读取第一行的数据，作为vn的长度
                        ‘file’ 从文件的首行读取vn，和数据共享分隔符
                        直接指定vn_lst列表
        @param omitlines: 文件开始几行无效数据
        @param separator: 除空白字符外的 额外分割符
        @param column_expected: 期望的列数
        @return:
        """
        f = open(pathname, 'r')

        # 处理跳过行
        if omitlines != 0:
            for i in range(omitlines):
                f.readline()

        # vnline=None
        # fisrtline=None
        # if vn_style=='file':#从文件中读取
        #     vnline=f.readline()
        # if vn_style=='numbers' and column_expected is None:#当vn模式为自动数字且不知道期望列数 才会读第一个数据行
        #     fisrtline=f.readline()
        #
        #
        # # 处理column_expected
        # self_column=column_expected == None #标识是否是自己通过第一行数据赋值给column_expected
        # if column_expected is None:  # 如果没有指定期望的列数，就用第一行数据的个数作为期望数据
        #     firsteles = get_elements_from_line(first, additional_separator=separator, number_only=False)
        #     column_expected = len(firsteles)
        #
        # #处理vn
        # if vn_style=='numbers':
        #     t=range(column_expected)
        #     vn_lst=['vn'+str(x) for x in t]#用vn0 vn1 vn2作为变量名
        # elif vn_style=='file':#从文件中读取
        #     vnline = f.readline()
        #     vn_lst = get_elements_from_line(vnline, additional_separator=separator, number_only=False)
        # elif isinstance(vn_style,list):#直接指定了vn_lst
        #     vn_lst=vn_style
        # else:
        #     raise Exception("参数错误")

        #
        self_column = False
        if vn_style == 'numbers' and column_expected is None:  # 需要读第一行数据
            self_column = True
            firstline = f.readline()
            firsteles = get_elements_from_line(firstline, additional_separator=separator, number_only=False)
            column_expected = len(firsteles)
            t = range(column_expected)
            vn_lst = ['vn' + str(x) for x in t]  # 用vn0 vn1 vn2作为变量名
        elif vn_style == 'numbers' and column_expected is not None:  # 不需要读第一行数据
            t = range(column_expected)
            vn_lst = ['vn' + str(x) for x in t]  # 用vn0 vn1 vn2作为变量名
        elif vn_style == 'file':  # 从文件中读取 此时不管column_expected
            vnline = f.readline()
            vn_lst = get_elements_from_line(vnline, additional_separator=separator, number_only=False)
            column_expected = len(vn_lst)
        elif isinstance(vn_style, list):  # 直接指定了vn_lst
            vn_lst = vn_style
            column_expected = len(vn_lst)
        else:
            raise Exception("参数错误")

        fdm = FlatDataModel()
        fdm.vn = vn_lst

        # 装载第一个数据
        if self_column:
            firstu = DataUnit()  # 装载第一个u
            for k, v in zip(vn_lst, firsteles):
                firstu.data[k] = v
            fdm.units.append(firstu)
            firstu.model = fdm

        # 开始读
        line = f.readline()
        while line:
            eles = get_elements_from_line(line, additional_separator=separator, number_only=False)
            u = DataUnit()  # 装载第一个u
            u.model = fdm
            for k, v in zip(vn_lst, eles):
                u.data[k] = v
            fdm.units.append(u)
            line = f.readline()
        f.close()

        return fdm

    def sort(self,key,reverse=False):
        """
        排序
        @param key:字段名
        @param reverse: 降序为true
        @return:
        """
        lams=[]
        if isinstance(key,str):
            key=[key]
        for i in key:
            assert i in self.vn,"不存在字段%s"%i
            lams.append(lambda x:x.data[i])#从字段名转换为取值的匿名函数
        func=lambda x:tuple([i(x) for i in lams])#这一段是取任意字段组成tuple的匿名函数
        self.units.sort(key=func,reverse=reverse)

    def make_bunches(self,classify_names)->Dict:
        """
        分类
        不会删除或者增加数据点
        @param classify_names: 用于分类的字段列表
        @return:
        """
        def make_legend_names():
            #生成图例名称
            nonlocal  classify_names,classify_names_value_unique
            r=[]
            for i in classify_names_value_unique:
                t = []
                for name, value in zip(classify_names, i):
                    t.append("%s:%s" % (name, value.__str__()))
                r.append(",".join(t))  # 添加图例名称 逗号分隔
            return r
        r={}
        if isinstance(classify_names,str):
            classify_names=[classify_names]
        self.sort(classify_names)
        classify_names_value=self[classify_names]#分类名的值
        classify_names_value_unique, firstid = myunique(classify_names_value, True)
        legend_names=make_legend_names()
        for i,name in enumerate(classify_names_value_unique):
            if i != len(classify_names_value_unique)-1:
                datas=self[firstid[i]:firstid[i+1]]
            else:
                datas=self[firstid[i]:]
            r[legend_names[i]]=self.__load_from_datas(self.vn,datas)
        return r

    def __load_from_datas(self,vnlist,datas)->'FlatDataModel':
        """
        从datas中生成fdm
        不用deepcopy
        @param vnlist:
        @param datas:
        @return:
        """
        assert is_sequence_with_specified_type(datas,DataUnit),'datas必须为dataunit组成的列表'
        fdm=FlatDataModel()
        fdm.vn=vnlist
        fdm.units=datas#这里没有deepcopy
        return fdm

    @staticmethod
    def load_from_string(stringtxt:str,vn_syle='fromstring',separator=' '):
        stringtxt=stringtxt.strip('\n ')#去除多余的换行符
        lns=stringtxt.split('\n')#按行划分
        fdm=FlatDataModel()
        if vn_syle == 'fromstring':#使用第一行的数据
            eles = get_elements_from_line(lns[0], additional_separator=separator, number_only=False)
            fdm.vn=eles
            lns.pop(0)#去除第一个list
        elif isinstance(vn_syle,list):
            fdm.vn=vn_syle
        else:
            raise TypeError("vn_style参数错误")
        for ln in lns:
            u=DataUnit()
            u.model=fdm
            eles = get_elements_from_line(ln, additional_separator=separator, number_only=False)
            for k,v in zip(fdm.vn,eles):
                u.data[k]=v
            fdm.units.append(u)
        return fdm

    def append_unit(self,u:DataUnit):
        """
        从dataunit中添加
        @param u:
        @return:
        """
        assert isinstance(u,DataUnit),'参数类型错误'
        u=deepcopy(u)
        self.units.append(u)
        u.model=self

    def merge(self,primary:str,secondary:Union[str,list],func:callable=None)->None:
        """
        合并某部分字段名
        如果中途发生错误，可能获得一个错误的fdm，且无法撤回。建议操作前备份
        @param primary:主字段名
        @param secondary: 次字段名，合并完成后会删除
        @param func:
        @return:
        """
        assert primary in self.vn,"未知字段%s"%primary
        if func is None:
            func=lambda x,y:x+y#默认相加
        assert callable(func),'func必须为二元函数'
        if isinstance(secondary,str):
            secondary=[secondary]
        assert isinstance(secondary,list),'secodary(%s)参数错误,必须为str或者str列表'%secondary
        # fdm=deepcopy(self)#复制一份自己 并在复制里面进行操作
        fdm=self
        for i in secondary:
            assert i in self.vn,"未知字段%s"%i
            for u in fdm:#操作
                u.data[primary]=func(u.data[primary],u.data[i])
            fdm.delete_variable(i)
        # return fdm
        pass

class TestCase(unittest.TestCase):
    def test_append_unit(self):
        vnlst = ['姓名', '性别', '列表']
        datalst = [['迈克尔', '男', 1], ['丹妮', '女', 2]]
        fdm1 = FlatDataModel.load_from_list(vnlst, datalst)
        fdm=FlatDataModel()
        fdm.vn=['姓名', '性别', '列表']
        fdm.append_unit(fdm1[0])
        self.assertEqual(1,fdm[0]['列表'])
        t=fdm1[0]
        t.data['列表']=2
        self.assertEqual(1, fdm[0]['列表'])#不跟着这个dataunit动
        # fdm = FlatDataModel()
        # fdm.vn = ['姓名', '性别', '列表']
        # fdm.append_unit(fdm1[0],flag_deepcopy=True)
        # self.assertEqual(2, len(fdm[0]['列表']))
        # t = fdm1[0]
        # t.data['列表'] = []
        # self.assertEqual(0, len(fdm[0]['列表']))#不会跟着这个dataunit动
    def test_load_from_list(self):
        vnlst = ['姓名', '性别', '年龄']
        datalst = [['迈克尔', '男', 1], ['丹妮', '女', 3]]
        fdm = FlatDataModel.load_from_list(vnlst, datalst)
        # print(fdm)
        # assert len(fdm.units) == 2
        # assert len(fdm.vn) == 3
        # assert fdm.units[0]['年龄'] == 1
        # assert fdm.units[1]['性别'] == '女'
        self.assertAlmostEqual(len(fdm.units), 2)
        self.assertAlmostEqual(len(fdm.vn), 3)
        self.assertAlmostEqual(fdm.units[0]['年龄'], 1)
        self.assertEqual(fdm.units[1]['性别'], '女')

    def test_load_from_file(self):
        fdm = FlatDataModel.load_from_file(pathname=r"test_load_from_file.txt",
                                           vn_style='numbers',
                                           omitlines=1,
                                           separator=','
                                           )
        print(fdm)
        self.assertEqual(len(fdm.units), 53)
        self.assertEqual(len(fdm.vn), 8)
        self.assertAlmostEqual(fdm.units[-2]['vn7'], -113001.148438)

    def test_load_from_file1(self):
        fdm = FlatDataModel.load_from_file(pathname=r"test_load_from_file.txt",
                                           vn_style=['1', '2', '3', '4', '5', '6', '7', '8'],
                                           omitlines=1,
                                           separator=','

                                           )
        print(fdm)
        self.assertEqual(len(fdm.units), 53)
        self.assertEqual(len(fdm.vn), 8)
        self.assertAlmostEqual(fdm.units[-2]['8'], -113001.148438)

    def test_load_from_file2(self):
        fdm = FlatDataModel.load_from_file(pathname=r"test_load_from_file.txt",
                                           vn_style='file',
                                           omitlines=0,
                                           separator=','

                                           )
        print(fdm)
        self.assertEqual(len(fdm.units), 53)
        self.assertEqual(len(fdm.vn), 8)
        self.assertAlmostEqual(fdm.units[-2]['data2'], -113001.148438)

    def test_flhz(self):  # 测试flhz
        fdm = FlatDataModel.load_from_excel_file(r"E:\我的文档\python\GoodToolPython\excel\test1.xlsx", 'Sheet1')
        o = fdm.flhz("拉杆屈服强度", [["P1底轴力", lambda x: len(x)]])
        self.assertEqual(42, o[0].data["P1底轴力"])
        self.assertEqual(28, o[1].data["P1底轴力"])
        self.assertEqual(28, o[-1].data["P1底轴力"])

        o = fdm.flhz("拉杆屈服强度", [["P1底轴力", lambda x: absmax(x)], ["P1底弯矩", lambda x: absmax(x), "弯矩"],
                                ["P1底弯矩", lambda x: abs(absmax(x)), "弯矩2"]])
        self.assertAlmostEqual(3682, o[0].data["P1底轴力"], delta=1)
        self.assertAlmostEqual(-11129, o[0].data["弯矩"], delta=1)
        self.assertAlmostEqual(11129, o[0].data["弯矩2"], delta=1)

    def test_getitem_setitem(self):
        vnlst = ['姓名', '性别', '年龄']
        datalst = [['迈克尔', '男', 1], ['丹妮', '女', 3],['雪落','男',11]]
        fdm = FlatDataModel.load_from_list(vnlst, datalst)
        print(fdm['姓名'])
        self.assertEqual(['迈克尔', '丹妮', '雪落'],fdm['姓名'])
        fdm['年龄2']=lambda x:-1*x['年龄']
        self.assertEqual([-1, -3, -11],fdm['年龄2'])
        fdm['年龄2'] = lambda x: 10+ x['年龄']
        print(fdm['年龄2'])
        self.assertEqual([11, 13, 21], fdm['年龄2'])
        fdm['年龄2'] =[1,1,1]
        self.assertEqual([1,1,1], fdm['年龄2'])


    def test_sort(self):
        vnlst = ['姓名', '性别', '年龄','身高']
        datalst = [['迈克尔', '男', 4,2], ['丹妮', '女', 3,2],['雪落','男',11,3]]
        fdm = FlatDataModel.load_from_list(vnlst, datalst)
        fdm.sort(key=['身高','年龄'])
        self.assertEquals('丹妮',fdm[0]['姓名'])

    def test_makebunches(self):
        vnlst = ['姓名', '性别', '年龄','身高']
        datalst = [['迈克尔', '男', 4,2], ['丹妮', '女', 3,2],['雪落','男',11,3],['卓哥','男',11,3.5]]
        fdm = FlatDataModel.load_from_list(vnlst, datalst)
        r=fdm.make_bunches(['性别','年龄'])
        self.assertEquals(3,len(r))
        self.assertEquals(2,len(r['性别:男,年龄:11']))
        pass

    def test_loadfromstring(self):
        st = """拉杆刚度	拉杆屈服强度	P1底轴力	P1底剪力
        1	50	3678.027091	570.7248771
        70000	50	3682.18235	558.0163549
        70000	100	3725.238945	554.0071293
        70000	150	3763.116878	551.450626
        70000	200	3795.721789	550.3307469
        70000	250	3821.993015	549.582025
        70000	300	3851.3423	548.9927673
        90000	50	3664.146017	557.0739193
        90000	100	3701.476992	552.4944948
        90000	150	3733.630435	548.1569784
        90000	200	3759.021624	546.8723534
        90000	250	3778.943963	545.8375741
        90000	300	3801.283942	544.9070614
        """
        fdm=FlatDataModel.load_from_string(st,separator='\t')
        # fdm.show_in_excel()
        self.assertEqual(4,len(fdm.vn))
        self.assertEqual(13, len(fdm))
        self.assertEqual(1,fdm[0]['拉杆刚度'])

        st = """
        1	50	3678.027091	570.7248771
        70000	50	3682.18235	558.0163549
        70000	100	3725.238945	554.0071293
        70000	150	3763.116878	551.450626
        70000	200	3795.721789	550.3307469
        70000	250	3821.993015	549.582025
        70000	300	3851.3423	548.9927673
        90000	50	3664.146017	557.0739193
        90000	100	3701.476992	552.4944948
        90000	150	3733.630435	548.1569784
        90000	200	3759.021624	546.8723534
        90000	250	3778.943963	545.8375741
        90000	300	3801.283942	544.9070614
        """
        fdm=FlatDataModel.load_from_string(st,vn_syle=['1','2','3','4'],separator='\t')
        self.assertEquals(4,len(fdm.vn))
        self.assertEquals(13, len(fdm))
        self.assertAlmostEqual(70000, fdm[2]['1'],delta=0.1)

    def test_readxls_with_mergedcells(self):
        fdm=FlatDataModel.load_from_excel_file(r"E:\我的文档\python\GoodToolPython\excel\test_合并单元格.xlsx")
        self.assertEqual("跨绕城高速大桥",fdm.units[2]['桥梁'])
        self.assertEqual("跨绕城高速大桥", fdm.units[3]['桥梁'])
        self.assertAlmostEqual(2031, fdm.units[3]['全长'],delta=1)


    def test_merge(self):
        vnlst = ['姓名', '性别', '年龄','身高']
        datalst = [['迈克尔', '男', 4,2], ['丹妮', '女', 3,2],['雪落','男',11,3],['卓哥','男',11,3.5]]
        fdm = FlatDataModel.load_from_list(vnlst, datalst)
        fdm.merge('年龄','身高')
        self.assertEqual(6,fdm[0]['年龄'])
        self.assertEqual(14.5, fdm[-1]['年龄'])

    def test_eff(self):#测试二分法
        vnlst = ['姓名', '性别', '年龄','身高']
        datalst = [['迈克尔', '男', 4,2], ['丹妮', '女', 3,2],['雪落','男',11,3],['卓哥','男',10,3.5]]
        fdm = FlatDataModel.load_from_list(vnlst, datalst)
        fdm.sort(key='年龄')
        i,lb,ub=bisection_method(sorted_list=fdm.units,
                                 goal=4,
                                 func=lambda x:x['年龄'])
        self.assertEqual(1,i)
        self.assertEqual(0, lb)
        i, lb, ub = bisection_method(sorted_list=fdm.units,
                                     goal=11,
                                     func=lambda x: x['年龄'])
        self.assertEqual(3, i)
        self.assertEqual(3, ub)
        i, lb, ub = bisection_method(sorted_list=fdm.units,
                                     goal=3,
                                     func=lambda x: x['年龄'])
        self.assertEqual(0, i)
        self.assertEqual(0, lb)
        i, lb, ub = bisection_method(sorted_list=fdm.units,
                                     goal=3.5,
                                     func=lambda x: x['年龄'])
        self.assertEqual(0, lb)
        self.assertEqual(1, ub)
        i, lb, ub = bisection_method(sorted_list=fdm.units,
                                     goal=10.01,
                                     func=lambda x: x['年龄'])
        self.assertEqual(2, lb)
        self.assertEqual(3, ub)
        self.assertRaises(Exception,bisection_method,sorted_list=fdm.units,
                                     goal=2,
                                     func=lambda x: x['年龄'])
if __name__ == '__main__':
    unittest.main()













    # fdm = FlatDataModel.load_from_excel_file(r"E:\我的文档\python\GoodToolPython\excel\OnLbvnclassChar.xlsx", 'Sheet1')
    # fdm.show_in_excel()

    # test_load_from_list()
    # fullname = "OnLbvnclassChar.xlsx"
    # # fullname1 = "D:\新建 Microsoft Excel 工作表.xlsx"
    # model = FlatDataModel.load_from_excel_file(fullname=fullname)
    # u = model[0]
    # print(u['文件名', '间距'])
    # print(model)
    # print(model['文件名'])

    # model.flhz(classify_names=['工况名','拉杆刚度'],
    #            statistics_func=[['P1底剪力',max,'p1底剪力'],
    #                             ['P1底剪力',len,'个数']])
    # model.flhz(classify_names=['工况名','拉杆刚度'],
    #            statistics_func=[['P1底剪力',max],
    #                             ]).append_to_file(fullname,'flhz')
